#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Full LVE pipeline (Option A, merged):
- Load up to 4 CSV/XLSX via GUI
- Clean & normalize faculty labels (remove ArchivXX / leading numbers; normalize 'Lehreinheit ...')
- Build:
    1) Faculty_Presentables.xlsx per semester (with Uni-vs-Fakultät line charts; no cumulative % in charts)
    2) Faculty_Over_Time.xlsx (multi-semester blocks + bar charts)
    3) Semester_Over_Time.xlsx (Uni-wide distributions per semester + charts)
    4) Semester_Faculty_Stats.xlsx (vertical semester blocks, Mittelwert colored, chart with faculties on X and semesters as series)
- Color thresholds for mean values: <61 red, <81 yellow, >=81 green
- Semester ordering: '23/24' < '24' < '24/25' and SS < '' < WS on ties
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Tuple
import re
import os
import pandas as pd

# GUI
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog
except Exception:
    tk = None

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ---------------- Configuration ----------------
BASE_REPORTS = Path.home() / "Desktop" / "Reports"
OUT_PRESENTABLES = "Faculty_Presentables.xlsx"
OUT_FAC_OVER_TIME = "Faculty_Over_Time.xlsx"
OUT_SEM_OVER_TIME = "Semester_Over_Time.xlsx"
OUT_SEM_FAC_STATS = "Semester_Faculty_Stats.xlsx"
OUTFILE_NAME = OUT_SEM_FAC_STATS  # keep compatibility

# ----------------- Robust CSV reading -----------------
def _read_csv_safely(path: str) -> pd.DataFrame:
    """Try common encodings & auto separators for CSV."""
    attempts = [
        {"encoding": "utf-8-sig", "sep": None, "engine": "python"},
        {"encoding": "cp1252",    "sep": None, "engine": "python"},
        {"encoding": "latin-1",   "sep": None, "engine": "python"},
    ]
    for kw in attempts:
        try:
            return pd.read_csv(path, **kw)
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
    # Last resort
    return pd.read_csv(
        path,
        encoding="latin-1",
        sep=None,
        engine="python",
        on_bad_lines="skip",
        encoding_errors="replace",
    )

# ----------------- UI helpers ------------------
def upload_file_for(name: str):
    """Prompt for a single file; allow skipping."""
    if tk is None:
        return None
    while True:
        file_path = filedialog.askopenfilename(
            title=f"Select data file for {name}",
            filetypes=(("CSV Files", "*.csv"),
                       ("Excel Files", "*.xlsx *.xls"),
                       ("All Files", "*.*"))
        )
        if not file_path:
            ans = simpledialog.askstring("Skip File?",
                                         f"No file selected for {name}.\nEnter Y to skip or N to choose again:")
            if ans and ans.strip().upper() == "Y":
                print(f"Skipping {name}.")
                return None
            else:
                continue
        else:
            ext = Path(file_path).suffix.lower()
            try:
                if ext == ".csv":
                    df_temp = _read_csv_safely(file_path)
                elif ext in (".xlsx", ".xls"):
                    df_temp = pd.read_excel(file_path)
                else:
                    messagebox.showerror("Unsupported File", "Please select a CSV or Excel file.")
                    continue
                messagebox.showinfo("Loaded", f"File for {name} loaded.")
                return df_temp
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read file.\n\n{e}")
                continue
# ---------------- Utilities / Helpers ----------------
def sanitize_sheet_name(name: str) -> str:
    """Excel-safe sheet name (<=31 chars; no []:*?/\\)."""
    safe = re.sub(r'[\[\]\*\?\/\\:]', '_', str(name)).strip()
    return safe[:31] or "Sheet"

# Semester ordering
def semester_sort_key(s: str) -> Tuple[int, int, int]:
    """
    Sort key so '23/24' < '24' < '24/25' and SS < '' < WS on ties.
    Returns (start, end, term_rank).
    """
    s = (s or "").strip()
    term_match = re.search(r'\b(WS|SS)\b', s, flags=re.IGNORECASE)
    term = term_match.group(1).upper() if term_match else ''
    m = re.search(r'(\d{2})(?:\s*/\s*(\d{2}))?$', s)
    if not m:
        return (99, 99, 2)
    start = int(m.group(1))
    end = int(m.group(2)) if m.group(2) else start
    term_rank = {'SS': 0, '': 1, 'WS': 2}.get(term, 1)
    return (start, end, term_rank)

STAT_ORDER = ["Anzahl", "Mittelwert", "Minimum", "Maximum", "Median", "Modus"]

def _stats(series: pd.Series) -> Dict[str, Optional[float]]:
    s = pd.to_numeric(series, errors="coerce").dropna()
    if s.empty:
        return dict(Anzahl=0, Mittelwert=None, Minimum=None, Maximum=None, Median=None, Modus=None)
    mode_series = s.mode()
    mode_val = float(mode_series.iloc[0]) if not mode_series.empty else None
    return dict(
        Anzahl=int(s.shape[0]),
        Mittelwert=round(float(s.mean()), 2),
        Minimum=round(float(s.min()), 2),
        Maximum=round(float(s.max()), 2),
        Median=round(float(s.median()), 2),
        Modus=round(mode_val, 2) if mode_val is not None else None,
    )

# Styling helpers
FILL_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # <61
FILL_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # <81
FILL_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # >=81
FILL_GREY   = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
THIN = Side(style="thin", color="000000")

def _apply_grid(ws, top: int, left: int, bottom: int, right: int) -> None:
    for r in range(top, bottom + 1):
        for c in range(left, right + 1):
            ws.cell(row=r, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _autosize(ws) -> None:
    for c in range(1, ws.max_column + 1):
        width = 12
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            width = max(width, min(60, len(str(v)) + 2))
        ws.column_dimensions[get_column_letter(c)].width = width

def _fill_mw_cell(cell, val: Optional[float]) -> None:
    if val is None:
        return
    if val < 61:
        cell.fill = FILL_RED
    elif val < 81:
        cell.fill = FILL_YELLOW
    else:
        cell.fill = FILL_GREEN

def _freeze_and_polish(ws):
    ws.freeze_panes = "B2"

# Chart helpers
def _set_series_name(series, name: str) -> None:
    lbl = SeriesLabel()
    lbl.v = str(name)
    series.tx = lbl

def _anchor_for_chart(ws) -> str:
    last_col = ws.max_column
    last_row = ws.max_row
    return "K3" if last_col >= 10 else f"A{last_row + 3}"

def _find_last_row(ws) -> int:
    r = ws.max_row
    while r > 1 and (
        ws.cell(row=r, column=1).value is None
        or str(ws.cell(row=r, column=1).value).strip() == ""
    ):
        r -= 1
    return r

def _collect_semesters_and_faculties_from_sheet(ws) -> Tuple[List[str], List[str], List[Tuple[str, int]]]:
    faculties: List[str] = []
    semesters: List[str] = []
    header_rows: List[Tuple[str, int]] = []
    r = 1
    maxr = ws.max_row
    while r <= maxr:
        a = ws.cell(row=r, column=1).value
        a_next = ws.cell(row=r + 1, column=1).value if r + 1 <= maxr else None
        if a and isinstance(a, str) and a.strip() and (a_next == "Anzahl"):
            sem_label = a.strip()
            header_rows.append((sem_label, r))
            semesters.append(sem_label)
            if not faculties:
                c = 2
                while True:
                    v = ws.cell(row=r, column=c).value
                    if v is None or str(v).strip() == "":
                        break
                    faculties.append(str(v).strip())
                    c += 1
        r += 1
    return faculties, semesters, header_rows
# --------- Faculty label cleaning helpers (extended) ---------
_ARCHIV_PREFIX_RE = re.compile(r'^\s*Archiv\s*\d{0,2}\s*', re.IGNORECASE)
_LEADING_NUM_RE   = re.compile(r'^\s*\d+\s*[.\-–:]*\s*')

def _de_umlaut_fix(s: str) -> str:
    # Normalize common ASCII fallbacks to German diacritics
    return (s.replace("Fakultat", "Fakultät")
             .replace("fakultat", "fakultät")
             .replace("fur", "für")
             .replace("Fur", "Für"))

def _canonical_parent_phrase(rest: str) -> str:
    """
    If trailing part indicates the known parent faculty (Mathematik & Informatik),
    return the canonical form; otherwise a cleaned version.
    """
    low = rest.lower()
    if "mathemat" in low and "informat" in low:
        return "Fakultät für Mathematik und Informatik"
    rest = rest.strip().strip("()[]")
    rest = _de_umlaut_fix(rest)
    if rest.lower().startswith("fakultat") or rest.lower().startswith("fakultät"):
        if not rest.lower().startswith("fakultät"):
            rest = rest.replace("Fakultat", "Fakultät").replace("fakultat", "fakultät")
    return rest

_LEHREINHEIT_RE = re.compile(
    r'^\s*Lehreinheit\s+([A-Za-zÄÖÜäöüß\-\s]+?)\s*(?:[\(\-–—]\s*(.*))?$',
    re.IGNORECASE
)

def _normalize_lehreinheit(s: str) -> Optional[str]:
    """
    'Lehreinheit Informatik (Fakultät für Mathematik und Informatik'
      -> 'Informatik - Fakultät für Mathematik und Informatik'
    """
    m = _LEHREINHEIT_RE.match(s)
    if not m:
        return None
    unit = m.group(1) or ""
    rest = (m.group(2) or "").strip()
    unit = re.sub(r'\s+', ' ', unit).strip()
    parent = _canonical_parent_phrase(rest) if rest else ""
    return f"{unit} - {parent}" if parent else unit

def _clean_faculty_label(x: str) -> str:
    """
    - Remove 'ArchivXX' prefix
    - Remove leading numbers/separators ('01', '01.', '12 –', '03:')
    - Normalize 'Lehreinheit X (…)' → 'X - …'
    - Fix ASCII umlauts and compact spaces
    """
    s = str(x or "").strip()
    s = _ARCHIV_PREFIX_RE.sub("", s)
    s = _LEADING_NUM_RE.sub("", s)
    normalized = _normalize_lehreinheit(s)
    if normalized is not None:
        s = normalized
    s = _de_umlaut_fix(s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# --------- Column mapping + normalization ------
def _map_columns(df: pd.DataFrame) -> Dict[str, str]:
    """
    Map user columns to canonical: Fakultät, Semester, GZI.
    Accepts common alternates.
    """
    candidates = {
        "Fakultät": ["Fakultät", "Fakultaet", "Fakultat", "Teilbereich", "Department"],
        "Semester": ["Semester", "Periode", "Term"],
        "GZI": ["Gesamtzufriedenheitsindex", "Qualitätsindex", "Qualitaetsindex",
                "Qualitäts-Index", "Qualitaets-Index", "Score", "Score (%)"],
    }
    colmap: Dict[str, Optional[str]] = {"Fakultät": None, "Semester": None, "GZI": None}
    lower = {c.lower(): c for c in df.columns}
    for key, opts in candidates.items():
        for o in opts:
            if o.lower() in lower:
                colmap[key] = lower[o.lower()]
                break
    missing = [k for k, v in colmap.items() if v is None]
    if missing:
        raise KeyError(f"Missing required columns: {missing}. Found: {list(df.columns)}")
    return {k: str(v) for k, v in colmap.items()}

def _normalize_df(df: pd.DataFrame, colmap: Dict[str, str]) -> pd.DataFrame:
    fac, sem, gzi = colmap["Fakultät"], colmap["Semester"], colmap["GZI"]
    out = df.copy()
    out[fac] = out[fac].astype(str).map(_clean_faculty_label)
    out[sem] = out[sem].astype(str).str.strip()
    if out[gzi].dtype == object:
        out[gzi] = out[gzi].astype(str).str.replace(",", ".", regex=False)
    out[gzi] = pd.to_numeric(out[gzi], errors="coerce")
    out = out.dropna(subset=[fac, sem, gzi])
    return out

def _faculty_list(df: pd.DataFrame, fac_col: str) -> List[str]:
    return sorted(df[fac_col].dropna().astype(str).unique(), key=lambda x: x.lower())

def _semester_list(df: pd.DataFrame, sem_col: str) -> List[str]:
    return sorted(df[sem_col].dropna().astype(str).unique(), key=semester_sort_key)
# --------- Bins & helpers (for presentables) ---------
def _bin_edges_and_labels() -> Tuple[List[int], List[str]]:
    # Edges: [0, 6, 11, ..., 96, 101] (right=False -> [start, next_start))
    edges = [0]
    v = 6
    while v <= 101:
        edges.append(v)
        v += 5
    labels = [f"{edges[i]}-{edges[i+1]-1}" for i in range(len(edges)-1)]
    return edges, labels

def compute_bins_from_gzi(series: pd.Series) -> pd.Series:
    edges, labels = _bin_edges_and_labels()
    s = pd.to_numeric(series, errors="coerce")
    return pd.cut(s, bins=edges, right=False, labels=labels, include_lowest=True, ordered=True)

def summary_stats(series: pd.Series):
    s = pd.to_numeric(series, errors="coerce").dropna()
    if s.empty:
        return {"Gesamt": 0, "Anzahl": 0, "Mittelwert": None, "Minimum": None,
                "Maximum": None, "Median": None, "Modus": None}
    mode_series = s.mode()
    mode_val = float(mode_series.iloc[0]) if not mode_series.empty else None
    return {
        "Gesamt": int(s.shape[0]),
        "Anzahl": int(s.shape[0]),
        "Mittelwert": round(float(s.mean()), 2),
        "Minimum": round(float(s.min()), 2),
        "Maximum": round(float(s.max()), 2),
        "Median": round(float(s.median()), 2),
        "Modus": round(mode_val, 2) if mode_val is not None else None
    }

# --------- Color bin rows by label upper bound (restored) ---------
def _parse_upper_from_bin_label(label: str) -> Optional[float]:
    if not isinstance(label, str):
        return None
    t = label.strip()
    if "-" not in t:
        return None
    try:
        return float(t.split("-", 1)[1].strip())
    except Exception:
        return None

def color_bin_rows(ws: Worksheet, first_bin_row: int, last_bin_row: int,
                   label_col: int, color_cols_from: int, color_cols_to: int) -> None:
    """Color bin rows based on bin label thresholds; also greys 'Gesamt' rows."""
    light_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow    = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    row_grey  = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    for r in range(first_bin_row, last_bin_row + 1):
        lab = ws.cell(row=r, column=label_col).value
        if isinstance(lab, str) and lab.strip().lower() == "gesamt":
            fill = row_grey
        else:
            ub = _parse_upper_from_bin_label(lab)
            if ub is None:
                continue
            fill = light_red if ub < 61 else yellow if ub < 81 else green
        for c in range(color_cols_from, color_cols_to + 1):
            ws.cell(row=r, column=c).fill = fill

# ---------- Bin tables for a single semester ----------
def faculty_bin_table(df_sem: pd.DataFrame, faculty_col: str, gzi_col: str, faculty: str) -> Dict[str, Dict[str, float]]:
    _, bins = _bin_edges_and_labels()
    sub = df_sem[df_sem[faculty_col] == faculty].copy()
    sub["bin_label"] = compute_bins_from_gzi(sub[gzi_col])
    freq = sub["bin_label"].value_counts().reindex(bins, fill_value=0)
    total = int(freq.sum())
    pct = (freq / total * 100.0).round(2) if total > 0 else pd.Series([0.0]*len(bins), index=bins)
    cum = pct.cumsum().round(2)
    return {b: {"Häufigkeit": int(freq[b]), "Prozente": float(pct[b]), "Kumulierte Prozente": float(cum[b])} for b in bins}

def uni_bin_table(df_sem: pd.DataFrame, gzi_col: str) -> Dict[str, Dict[str, float]]:
    _, bins = _bin_edges_and_labels()
    ds = df_sem.copy()
    ds["bin_label"] = compute_bins_from_gzi(ds[gzi_col])
    freq = ds["bin_label"].value_counts().reindex(bins, fill_value=0)
    total = int(freq.sum())
    pct = (freq / total * 100.0).round(2) if total > 0 else pd.Series([0.0]*len(bins), index=bins)
    cum = pct.cumsum().round(2)
    return {b: {"Uni Häufigkeit": int(freq[b]), "Uni Prozente": float(pct[b]), "Uni Kumulierte Prozente": float(cum[b])} for b in bins}
# ---------- Per-semester workbook: one sheet per faculty ----------
def build_faculty_presentables_for_semester(df: pd.DataFrame, semester_value: str, out_folder: Path) -> Path:
    """
    Creates {out_folder}/Faculty_Presentables.xlsx with one sheet per faculty:
      - Binned table (Häufigkeit, Prozente, Kumulierte Prozente) for faculty + Uni
      - Semester label preserved in header
      - Two line charts (Uni vs Fakultät): Häufigkeit & Prozente across bins
    """
    out_folder.mkdir(parents=True, exist_ok=True)
    out_path = out_folder / OUT_PRESENTABLES

    # Map columns
    m = _map_columns(df)
    fac_col, sem_col, gzi_col = m["Fakultät"], m["Semester"], m["GZI"]

    # Normalize + filter to the semester
    df2 = _normalize_df(df, m)
    df_sem = df2[df2[sem_col] == str(semester_value)].copy()
    faculties = sorted(df_sem[fac_col].dropna().unique(), key=lambda x: str(x))
    _, bin_labels = _bin_edges_and_labels()

    uni_bin = uni_bin_table(df_sem, gzi_col)
    uni_sum = summary_stats(df_sem[gzi_col])

    wb = Workbook()
    ws_index = wb.active
    ws_index.title = "Index"
    ws_index.cell(row=1, column=1, value="Fakultät").font = Font(bold=True)
    ws_index.cell(row=1, column=2, value="Semester").font = Font(bold=True)

    grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    row_index = 2

    for fac in faculties:
        sheet = sanitize_sheet_name(str(fac)) or "Unbenannt"
        ws = wb.create_sheet(title=sheet)

        # Header row: faculty name + merged semester label
        ws.cell(row=1, column=1, value=str(fac)).font = Font(bold=True)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
        top = ws.cell(row=1, column=2)
        top.value = str(semester_value); top.alignment = center; top.font = Font(bold=True); top.fill = grey

        # Column headers
        ws.cell(row=2, column=1, value="Gesamtzufriedenheitsindex").alignment = center
        ws.cell(row=2, column=1).font = Font(bold=True); ws.cell(row=2, column=1).fill = grey
        fac_headers = ["Häufigkeit", "Prozente", "Kumulierte Prozente"]  # B..D
        uni_headers = ["Uni Häufigkeit", "Uni Prozente", "Uni Kumulierte Prozente"]  # E..G
        for j, lab in enumerate(fac_headers, start=2):
            c = ws.cell(row=2, column=j, value=lab); c.font = Font(bold=True); c.alignment = center; c.fill = grey
        for j, lab in enumerate(uni_headers, start=5):
            c = ws.cell(row=2, column=j, value=lab); c.font = Font(bold=True); c.alignment = center; c.fill = grey

        # Faculty distribution + summary
        fac_bin = faculty_bin_table(df_sem, fac_col, gzi_col, fac)
        fac_sum = summary_stats(df_sem.loc[df_sem[fac_col] == fac, gzi_col])

        # Bin rows
        first_bin_row = 3
        cur = first_bin_row
        for b in bin_labels:
            ws.cell(row=cur, column=1, value=b)
            fvals = fac_bin[b]; uvals = uni_bin[b]
            ws.cell(row=cur, column=2, value=fvals["Häufigkeit"])
            ws.cell(row=cur, column=3, value=fvals["Prozente"])
            ws.cell(row=cur, column=4, value=fvals["Kumulierte Prozente"])
            ws.cell(row=cur, column=5, value=uvals["Uni Häufigkeit"])
            ws.cell(row=cur, column=6, value=uvals["Uni Prozente"])
            ws.cell(row=cur, column=7, value=uvals["Uni Kumulierte Prozente"])
            cur += 1

        # Summary rows
        summary_labels = ["Gesamt", "Anzahl", "Mittelwert", "Minimum", "Maximum", "Median", "Modus"]
        first_summary_row = cur
        for i, label in enumerate(summary_labels):
            rr = first_summary_row + i
            ws.cell(row=rr, column=1, value=label).font = Font(bold=True)
            ws.cell(row=rr, column=2, value=fac_sum[label])
            ws.cell(row=rr, column=5, value=uni_sum[label])
            ws.cell(row=rr, column=2).number_format = "0" if label in {"Gesamt", "Anzahl"} else "0.00"
            ws.cell(row=rr, column=5).number_format = "0" if label in {"Gesamt", "Anzahl"} else "0.00"

        last_data_row = first_summary_row + len(summary_labels) - 1

        # Borders + autosize
        THIN = Side(style="thin", color="000000")
        for r in range(1, last_data_row + 1):
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        for c in range(1, ws.max_column + 1):
            width = 12
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    width = max(width, min(60, len(str(v)) + 2))
            ws.column_dimensions[get_column_letter(c)].width = width

        ws.freeze_panes = "B3"

        # ---- Uni vs Fakultät line charts (Häufigkeit & Prozente) ----
        last_bin_row = first_summary_row - 1
        if last_bin_row >= first_bin_row:
            cats = Reference(ws, min_col=1, min_row=first_bin_row, max_row=last_bin_row)

            # 1) Häufigkeit
            lc_h = LineChart()
            lc_h.title = "Uni vs Fakultät – Häufigkeit"
            lc_h.y_axis.title = "Häufigkeit"
            lc_h.x_axis.title = "GZI-Klassen"
            lc_h.legend.position = "b"; lc_h.width = 28; lc_h.height = 12
            lc_h.set_categories(cats)

            s_fac_h = Reference(ws, min_col=2, min_row=first_bin_row, max_row=last_bin_row)
            s_uni_h = Reference(ws, min_col=5, min_row=first_bin_row, max_row=last_bin_row)
            lc_h.add_data(s_fac_h, titles_from_data=False); _set_series_name(lc_h.series[-1], "Fakultät – Häufigkeit")
            lc_h.add_data(s_uni_h, titles_from_data=False); _set_series_name(lc_h.series[-1], "Uni – Häufigkeit")

            # 2) Prozente
            lc_p = LineChart()
            lc_p.title = "Uni vs Fakultät – Prozente"
            lc_p.y_axis.title = "Prozente (%)"
            lc_p.x_axis.title = "GZI-Klassen"
            lc_p.legend.position = "b"; lc_p.width = 28; lc_p.height = 12
            lc_p.set_categories(cats)

            s_fac_p = Reference(ws, min_col=3, min_row=first_bin_row, max_row=last_bin_row)
            s_uni_p = Reference(ws, min_col=6, min_row=first_bin_row, max_row=last_bin_row)
            lc_p.add_data(s_fac_p, titles_from_data=False); _set_series_name(lc_p.series[-1], "Fakultät – Prozente")
            lc_p.add_data(s_uni_p, titles_from_data=False); _set_series_name(lc_p.series[-1], "Uni – Prozente")

            # Place charts (right of table)
            ws.add_chart(lc_h, "I3")
            ws.add_chart(lc_p, "I20")

        # Index sheet link
        ws_index.cell(row=row_index, column=1, value=str(fac)).hyperlink = f"#{sheet}!A1"
        ws_index.cell(row=row_index, column=2, value=str(semester_value))
        row_index += 1

    # Style index a bit
    for c in (1, 2):
        ws_index.cell(row=1, column=c).fill = grey
        ws_index.cell(row=1, column=c).alignment = center
        ws_index.cell(row=1, column=c).font = Font(bold=True)

    wb.save(out_path)
    return out_path
# ---------- Faculty_Over_Time (multi-semester) ----------
def build_faculty_over_time_workbook(full_df: pd.DataFrame, base_reports_folder: Path) -> Path:
    out_path = base_reports_folder / OUT_FAC_OVER_TIME
    if full_df is None or full_df.empty:
        return out_path

    # Map + normalize
    colmap = _map_columns(full_df)
    fac_col, sem_col, gzi_col = colmap["Fakultät"], colmap["Semester"], colmap["GZI"]
    df = _normalize_df(full_df, colmap)

    faculties = sorted(df[fac_col].dropna().unique(), key=lambda x: str(x))
    semesters_all = sorted(df[sem_col].dropna().unique(), key=semester_sort_key)
    _, bin_labels = _bin_edges_and_labels()
    summary_labels = ["Gesamt", "Anzahl", "Mittelwert", "Minimum", "Maximum", "Median", "Modus"]

    # Precompute per (faculty, semester)
    facsem_bins: Dict[tuple, Dict[str, Dict[str, float]]] = {}
    facsem_summ: Dict[tuple, Dict[str, Optional[float]]] = {}

    for fac in faculties:
        dff = df[df[fac_col] == fac]
        for sem in semesters_all:
            dfs = dff[dff[sem_col] == sem].copy()
            dfs["bin_label"] = compute_bins_from_gzi(dfs[gzi_col])
            freq = dfs["bin_label"].value_counts().reindex(bin_labels, fill_value=0)
            total = int(freq.sum())
            pct = (freq / total * 100.0).round(2) if total > 0 else pd.Series([0.0]*len(bin_labels), index=bin_labels)
            cum = pct.cumsum().round(2)
            facsem_bins[(fac, sem)] = {b: {"Häufigkeit": int(freq[b]), "Prozente": float(pct[b]), "Kumulierte Prozente": float(cum[b])}
                                       for b in bin_labels}
            facsem_summ[(fac, sem)] = summary_stats(dfs[gzi_col])

    wb = Workbook()
    ws_index = wb.active
    ws_index.title = "Index"
    ws_index.cell(row=1, column=1, value="Fakultät").font = Font(bold=True)
    ws_index.cell(row=1, column=2, value="# Semester").font = Font(bold=True)

    grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    row_index = 2

    for fac in faculties:
        present_semesters = [s for s in semesters_all if not df[(df[fac_col]==fac) & (df[sem_col]==s)].empty]
        if not present_semesters:
            continue

        sheet_name = sanitize_sheet_name(str(fac)) or "Unbenannt"
        ws = wb.create_sheet(title=sheet_name)

        ws.cell(row=1, column=1, value=str(fac)).font = Font(bold=True)

        start_col = 2
        block_width = 3
        headers = ["Häufigkeit", "Prozente", "Kumulierte Prozente"]

        # Row 1: merged block captions (semester); Row 2: subheaders
        for i, sem in enumerate(present_semesters):
            blk_start = start_col + i*block_width
            blk_end = blk_start + block_width - 1
            ws.merge_cells(start_row=1, start_column=blk_start, end_row=1, end_column=blk_end)
            top = ws.cell(row=1, column=blk_start)
            top.value = sem; top.fill = grey; top.font = Font(bold=True); top.alignment = center
            for j, h in enumerate(headers):
                c = ws.cell(row=2, column=blk_start + j, value=h)
                c.fill = grey; c.font = Font(bold=True); c.alignment = center

        ws.cell(row=2, column=1, value="Gesamtzufriedenheitsindex").fill = grey
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=1).alignment = center

        first_bin_row = 3
        # Bin rows
        for r, b in enumerate(_bin_edges_and_labels()[1], start=first_bin_row):
            ws.cell(row=r, column=1, value=b)
            for i, sem in enumerate(present_semesters):
                blk_start = start_col + i*block_width
                dat = facsem_bins[(fac, sem)][b]
                ws.cell(row=r, column=blk_start + 0, value=dat["Häufigkeit"])
                ws.cell(row=r, column=blk_start + 1, value=dat["Prozente"])
                ws.cell(row=r, column=blk_start + 2, value=dat["Kumulierte Prozente"])

        # Summary rows (faculty-only)
        first_summary_row = first_bin_row + len(_bin_edges_and_labels()[1])
        for i, lab in enumerate(summary_labels):
            rr = first_summary_row + i
            ws.cell(row=rr, column=1, value=lab)
            for k, sem in enumerate(present_semesters):
                blk_start = start_col + k*block_width
                val = facsem_summ[(fac, sem)][lab]
                ws.cell(row=rr, column=blk_start + 0, value=val)
                ws.cell(row=rr, column=blk_start + 0).number_format = "0" if lab in {"Gesamt", "Anzahl"} else "0.00"

        last_row = first_summary_row + len(summary_labels) - 1
        last_col = start_col + len(present_semesters)*block_width - 1

        # Borders + separators + coloring
        THIN = Side(style="thin", color="000000")
        MED = Side(style="medium", color="000000")
        for r in range(1, last_row + 1):
            for c in range(1, last_col + 1):
                ws.cell(row=r, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        for i in range(len(present_semesters)):
            sep_col = start_col + i*block_width + (block_width - 1)
            for r in range(1, last_row + 1):
                cell = ws.cell(row=r, column=sep_col)
                cell.border = Border(left=cell.border.left, right=MED, top=cell.border.top, bottom=cell.border.bottom)

        color_bin_rows(ws, first_bin_row=first_bin_row, last_bin_row=first_summary_row - 1,
                       label_col=1, color_cols_from=1, color_cols_to=last_col)

        # Grey "Gesamt"
        row_grey = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        for rr in range(first_summary_row, last_row + 1):
            if str(ws.cell(row=rr, column=1).value).strip().lower() == "gesamt":
                for cc in range(1, last_col + 1):
                    ws.cell(row=rr, column=cc).fill = row_grey

        # Freeze & autosize
        ws.freeze_panes = "B3"
        for c in range(1, last_col + 1):
            width = 12
            for r in range(1, last_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    width = max(width, min(60, len(str(v)) + 2))
            ws.column_dimensions[get_column_letter(c)].width = width

        # ---- Charts (multi-semester) ----
        cats_ref = Reference(ws, min_col=1, min_row=first_bin_row, max_row=first_summary_row - 1)

        chart1 = BarChart(); chart1.type = "col"; chart1.grouping = "clustered"
        chart1.title = "GZI-Klassen - Häufigkeiten (Zeitvergleich)"
        chart1.y_axis.title = "Häufigkeiten"; chart1.x_axis.title = "Klassen"
        chart1.legend.position = "b"; chart1.width = 40; chart1.height = 18
        chart1.set_categories(cats_ref)

        chart2 = BarChart(); chart2.type = "col"; chart2.grouping = "clustered"
        chart2.title = "GZI-Klassen - Prozente (Zeitvergleich)"
        chart2.y_axis.title = "Prozente (%)"; chart2.x_axis.title = "Klassen"
        chart2.legend.position = "b"; chart2.width = 40; chart2.height = 18
        chart2.set_categories(cats_ref)

        for i, sem in enumerate(present_semesters):
            blk_start = start_col + i*block_width

            ref_h = Reference(ws, min_col=blk_start + 0, max_col=blk_start + 0,
                              min_row=first_bin_row, max_row=first_summary_row - 1)
            chart1.add_data(ref_h, titles_from_data=False)
            s1 = chart1.series[-1]; s1.tx = SeriesLabel(); s1.tx.v = str(sem)

            ref_p = Reference(ws, min_col=blk_start + 1, max_col=blk_start + 1,
                              min_row=first_bin_row, max_row=first_summary_row - 1)
            chart2.add_data(ref_p, titles_from_data=False)
            s2 = chart2.series[-1]; s2.tx = SeriesLabel(); s2.tx.v = str(sem)

        ws.add_chart(chart1, f"A{last_row + 3}")
        ws.add_chart(chart2, f"A{last_row + 25}")

        # Index entry
        ws_index.cell(row=row_index, column=1, value=str(fac)).hyperlink = f"#{sheet_name}!A1"
        ws_index.cell(row=row_index, column=2, value=len(present_semesters))
        row_index += 1

    # Style index + autosize
    for c in (1, 2):
        ws_index.cell(row=1, column=c).fill = grey
        ws_index.cell(row=1, column=c).alignment = center
    for c in range(1, ws_index.max_column + 1):
        width = 12
        for r in range(1, ws_index.max_row + 1):
            v = ws_index.cell(row=r, column=c).value
            if v is not None:
                width = max(width, min(60, len(str(v)) + 2))
        ws_index.column_dimensions[get_column_letter(c)].width = width

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(out_path)
    return out_path

# ---------- Semester_Over_Time (Uni over time) ----------
def build_semester_over_time_workbook(full_df: pd.DataFrame, base_reports_folder: Path, split_ws_ss: bool = False) -> Path:
    out_path = base_reports_folder / OUT_SEM_OVER_TIME
    if full_df is None or full_df.empty:
        return out_path

    # Column detection + normalize
    colmap = _map_columns(full_df)
    sem_col, gzi_col = colmap["Semester"], colmap["GZI"]
    df = _normalize_df(full_df, colmap)
    df[sem_col] = df[sem_col].astype(str)

    def build_sheet(ws: Worksheet, semesters: List[str], title: str):
        ws.title = sanitize_sheet_name(title)
        _, bin_labels = _bin_edges_and_labels()
        grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")

        # Row 1–2 headers
        ws.cell(row=2, column=1, value="Gesamtzufriedenheitsindex").fill = grey
        ws.cell(row=2, column=1).font = bold
        ws.cell(row=2, column=1).alignment = center

        start_col = 2
        block_width = 3
        sub_headers = ["Häufigkeit", "Prozente", "Kumulierte Prozente"]

        for i, sem in enumerate(semesters):
            blk_start = start_col + i * block_width
            blk_end = blk_start + block_width - 1
            ws.merge_cells(start_row=1, start_column=blk_start, end_row=1, end_column=blk_end)
            top = ws.cell(row=1, column=blk_start)
            top.value = sem
            top.fill = grey
            top.font = bold
            top.alignment = center
            for j, h in enumerate(sub_headers):
                c = ws.cell(row=2, column=blk_start + j, value=h)
                c.fill = grey
                c.font = bold
                c.alignment = center

        # Uni-wide distributions per semester
        first_bin_row = 3
        for r, b in enumerate(bin_labels, start=first_bin_row):
            ws.cell(row=r, column=1, value=b)

        # Compute per semester
        summary_labels = ["Gesamt", "Anzahl", "Mittelwert", "Minimum", "Maximum", "Median", "Modus"]
        first_summary_row = first_bin_row + len(bin_labels)

        for i, sem in enumerate(semesters):
            blk_start = start_col + i * block_width
            dss = df[df[sem_col] == sem].copy()
            dss["bin_label"] = compute_bins_from_gzi(dss[gzi_col])

            freq = dss["bin_label"].value_counts().reindex(bin_labels, fill_value=0)
            total = int(freq.sum())
            pct = (freq / total * 100.0).round(2) if total > 0 else pd.Series([0.0] * len(bin_labels), index=bin_labels)
            cum = pct.cumsum().round(2)

            for r, b in enumerate(bin_labels, start=first_bin_row):
                ws.cell(row=r, column=blk_start + 0, value=int(freq[b]))
                ws.cell(row=r, column=blk_start + 1, value=float(pct[b]))
                ws.cell(row=r, column=blk_start + 2, value=float(cum[b]))

            sm = summary_stats(dss[gzi_col])
            for k, lab in enumerate(summary_labels):
                rr = first_summary_row + k
                ws.cell(row=rr, column=1, value=lab)
                ws.cell(row=rr, column=blk_start + 0, value=sm[lab])
                ws.cell(row=rr, column=blk_start + 0).number_format = "0" if lab in {"Gesamt", "Anzahl"} else "0.00"

        last_row = first_summary_row + len(summary_labels) - 1
        last_col = start_col + len(semesters) * block_width - 1

        THIN = Side(style="thin", color="000000")
        MED = Side(style="medium", color="000000")
        for r in range(1, last_row + 1):
            for c in range(1, last_col + 1):
                ws.cell(row=r, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        for i in range(len(semesters)):
            sep_col = start_col + i * block_width + (block_width - 1)
            for r in range(1, last_row + 1):
                cell = ws.cell(row=r, column=sep_col)
                cell.border = Border(left=cell.border.left, right=MED, top=cell.border.top, bottom=cell.border.bottom)

        color_bin_rows(ws, first_bin_row=first_bin_row, last_bin_row=first_summary_row - 1,
                       label_col=1, color_cols_from=1, color_cols_to=last_col)

        row_grey = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        for rr in range(first_summary_row, last_row + 1):
            if str(ws.cell(row=rr, column=1).value).strip().lower() == "gesamt":
                for cc in range(1, last_col + 1):
                    ws.cell(row=rr, column=cc).fill = row_grey

        ws.freeze_panes = "B3"
        for c in range(1, last_col + 1):
            width = 12
            for r in range(1, last_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    width = max(width, min(60, len(str(v)) + 2))
            ws.column_dimensions[get_column_letter(c)].width = width

        cats_ref = Reference(ws, min_col=1, min_row=first_bin_row, max_row=first_summary_row - 1)

        ch1 = BarChart(); ch1.type = "col"; ch1.grouping = "clustered"
        ch1.title = "GZI-Klassen - Häufigkeiten (Zeitvergleich)"
        ch1.y_axis.title = "Häufigkeiten"; ch1.x_axis.title = "Klassen"
        ch1.legend.position = "b"; ch1.width = 40; ch1.height = 18
        ch1.set_categories(cats_ref)

        ch2 = BarChart(); ch2.type = "col"; ch2.grouping = "clustered"
        ch2.title = "GZI-Klassen - Prozente (Zeitvergleich)"
        ch2.y_axis.title = "Prozente (%)"; ch2.x_axis.title = "Klassen"
        ch2.legend.position = "b"; ch2.width = 40; ch2.height = 18
        ch2.set_categories(cats_ref)

        for i, sem in enumerate(semesters):
            blk_start = start_col + i * block_width

            ref_h = Reference(ws, min_col=blk_start + 0, max_col=blk_start + 0,
                              min_row=first_bin_row, max_row=first_summary_row - 1)
            ch1.add_data(ref_h, titles_from_data=False)
            s1 = ch1.series[-1]; s1.tx = SeriesLabel(); s1.tx.v = str(sem)

            ref_p = Reference(ws, min_col=blk_start + 1, max_col=blk_start + 1,
                              min_row=first_bin_row, max_row=first_summary_row - 1)
            ch2.add_data(ref_p, titles_from_data=False)
            s2 = ch2.series[-1]; s2.tx = SeriesLabel(); s2.tx.v = str(sem)

        ws.add_chart(ch1, f"A{last_row + 3}")
        ws.add_chart(ch2, f"A{last_row + 25}")

    wb = Workbook()
    ws_index = wb.active
    ws_index.title = "Index"
    ws_index.cell(row=1, column=1, value="Sheet").font = Font(bold=True)
    ws_index.cell(row=1, column=2, value="# Semester").font = Font(bold=True)

    semesters_all = sorted(df[sem_col].dropna().unique(), key=semester_sort_key)
    ws_all = wb.create_sheet(title="Uni_Over_Time")
    build_sheet(ws_all, semesters_all, "Uni_Over_Time")
    ws_index.cell(row=2, column=1, value="Uni_Over_Time").hyperlink = "#Uni_Over_Time!A1"
    ws_index.cell(row=2, column=2, value=len(semesters_all))

    row_idx = 3
    if split_ws_ss:
        ws_only = [s for s in semesters_all if re.search(r'\bWS\b', s, re.IGNORECASE)]
        ss_only = [s for s in semesters_all if re.search(r'\bSS\b', s, re.IGNORECASE)]
        if ws_only:
            ws_ws = wb.create_sheet(title="WS_Over_Time")
            build_sheet(ws_ws, ws_only, "WS_Over_Time")
            ws_index.cell(row=row_idx, column=1, value="WS_Over_Time").hyperlink = "#WS_Over_Time!A1"
            ws_index.cell(row=row_idx, column=2, value=len(ws_only)); row_idx += 1
        if ss_only:
            ws_ss = wb.create_sheet(title="SS_Over_Time")
            build_sheet(ws_ss, ss_only, "SS_Over_Time")
            ws_index.cell(row=row_idx, column=1, value="SS_Over_Time").hyperlink = "#SS_Over_Time!A1"
            ws_index.cell(row=row_idx, column=2, value=len(ss_only)); row_idx += 1

    grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    for c in (1, 2):
        ws_index.cell(row=1, column=c).fill = grey
        ws_index.cell(row=1, column=c).alignment = center
    for c in range(1, ws_index.max_column + 1):
        width = 12
        for r in range(1, ws_index.max_row + 1):
            v = ws_index.cell(row=r, column=c).value
            if v is not None:
                width = max(width, min(60, len(str(v)) + 2))
        ws_index.column_dimensions[get_column_letter(c)].width = width

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(out_path)
    return out_path
# ---------- Semester_Faculty_Stats: blocks with stats ----------
def build_semester_faculty_stats_blocks(full_df: pd.DataFrame, out_path: Path) -> Path:
    """
    Writes <Reports>/Semester_Faculty_Stats.xlsx with vertically stacked semester blocks.
    Layout per semester:
      • Header row: [A]=Semester, [B..]=Faculties (normalized)
      • Next 6 rows: Anzahl, Mittelwert, Minimum, Maximum, Median, Modus
      • Mittelwert row cells are color-coded: <61 red, <81 yellow, >=81 green
      • One blank spacer row between semesters
    """
    colmap = _map_columns(full_df)
    df = _normalize_df(full_df, colmap)

    fac_col, sem_col, gzi_col = colmap["Fakultät"], colmap["Semester"], colmap["GZI"]
    faculties = _faculty_list(df, fac_col)
    semesters = _semester_list(df, sem_col)

    wb = Workbook()
    ws = wb.active
    ws.title = "Semester_Faculty_Stats"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    row = 1
    for sem in semesters:
        # Header row
        ws.cell(row=row, column=1, value=sem).font = bold
        ws.cell(row=row, column=1).fill = FILL_GREY
        for j, fac in enumerate(faculties, start=2):
            c = ws.cell(row=row, column=j, value=fac)
            c.font = bold
            c.alignment = center
            c.fill = FILL_GREY

        # Stats rows
        dss = df[df[sem_col] == sem]
        for k, stat_name in enumerate(STAT_ORDER, start=1):
            r = row + k
            ws.cell(row=r, column=1, value=stat_name).font = bold
            for j, fac in enumerate(faculties, start=2):
                dssf = dss[dss[fac_col] == fac]
                stats = _stats(dssf[gzi_col])
                val = stats.get(stat_name)
                cell = ws.cell(row=r, column=j, value=val if val is not None else None)
                cell.number_format = "0" if stat_name == "Anzahl" else "0.00"
                if stat_name == "Mittelwert":
                    _fill_mw_cell(cell, val)

        # Borders around the semester block
        top = row
        bottom = row + len(STAT_ORDER)
        left = 1
        right = 1 + len(faculties)
        _apply_grid(ws, top, left, bottom, right)

        # Next block (leave one blank row)
        row = bottom + 2

    _freeze_and_polish(ws)
    _autosize(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path

# ---------- Chart matrix + chart (Faculties on X, Semesters as series) ----------
def append_chart_to_semester_faculty_stats(out_path: Path) -> Path:
    """
    Appends a hidden 'Chart_Data' sheet and a clustered column chart to the
    'Semester_Faculty_Stats' sheet. X-axis shows faculty names, one series per semester.
    """
    wb = load_workbook(out_path)
    ws = wb.active  # "Semester_Faculty_Stats"

    faculties, semesters, headers = _collect_semesters_and_faculties_from_sheet(ws)
    if not faculties or not semesters:
        wb.save(out_path)
        return out_path

    # Build/refresh compact Mittelwert matrix
    if "Chart_Data" in wb.sheetnames:
        del wb["Chart_Data"]
    wcd = wb.create_sheet("Chart_Data")

    # Header row: A1=Semester label, B..=faculties
    wcd.cell(row=1, column=1, value="Semester").font = Font(bold=True)
    for j, fac in enumerate(faculties, start=2):
        wcd.cell(row=1, column=j, value=fac).font = Font(bold=True)

    # Mittelwert row offset inside each semester block
    mw_offset = STAT_ORDER.index("Mittelwert") + 1  # Anzahl=1, Mittelwert=2, ...

    # Fill rows: one row per semester
    for i, (sem, r0) in enumerate(headers, start=2):
        wcd.cell(row=i, column=1, value=sem)
        r_mw = r0 + mw_offset
        for j in range(2, 2 + len(faculties)):
            val = ws.cell(row=r_mw, column=j).value
            try:
                val = float(val) if val is not None and str(val).strip() != "" else None
            except Exception:
                val = None
            cell = wcd.cell(row=i, column=j, value=val)
            cell.number_format = "0.00"

    # Detect bounds
    last_col = 1
    c = 2
    while True:
        v = wcd.cell(row=1, column=c).value
        if v is None or str(v).strip() == "":
            break
        last_col = c
        c += 1
    if last_col < 2:
        wb.save(out_path)
        return out_path

    last_row = 1
    r = 2
    while True:
        v = wcd.cell(row=r, column=1).value
        if v is None or str(v).strip() == "":
            break
        last_row = r
        r += 1
    if last_row < 2:
        wb.save(out_path)
        return out_path

    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Mittelwert je Fakultät (Semestergestaffelt)"
    chart.y_axis.title = "Mittelwert"
    chart.x_axis.title = "Fakultäten"
    chart.legend.position = "b"
    chart.width = 40
    chart.height = 18

    # X-axis = faculty headers
    cats = Reference(wcd, min_col=2, max_col=last_col, min_row=1, max_row=1)
    chart.set_categories(cats)

    # Data = rows (semesters) across faculties, series names from col A
    data_range = Reference(
        wcd,
        min_col=1,
        min_row=2,
        max_col=last_col,
        max_row=last_row
    )
    chart.add_data(data_range, titles_from_data=True, from_rows=True)

    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1

    ws.add_chart(chart, _anchor_for_chart(ws))
    wcd.sheet_state = "hidden"

    wb.save(out_path)
    return out_path

# ------------------ Public API (build + chart) ------------------
def build_semester_faculty_stats_v2(full_df: pd.DataFrame, base_reports_folder: Path) -> Path:
    out_path = base_reports_folder / OUTFILE_NAME
    out_path = build_semester_faculty_stats_blocks(full_df, out_path)
    out_path = append_chart_to_semester_faculty_stats(out_path)

    wb = load_workbook(out_path)
    ws = wb.active
    if ws.max_column >= 2 and ws.max_row >= 1:
        for c in range(2, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(out_path)
    return out_path
# ------------------- Main flow ------------------
def main():
    BASE_REPORTS.mkdir(parents=True, exist_ok=True)

    if tk is None:
        print("Tkinter GUI not available. Please run on a system with a display.")
        return

    root = tk.Tk()
    root.withdraw()

    # 1) Choose up to 4 datasets
    uploaded_frames: List[pd.DataFrame] = []
    pick_names = ["df", "df1", "df2", "df3"]
    for nm in pick_names:
        print(f"Select file for {nm}")
        tmp = upload_file_for(nm)
        if tmp is not None and not tmp.empty:
            uploaded_frames.append(tmp)

    if not uploaded_frames:
        messagebox.showwarning("No Data", "No datasets were selected. Exiting.")
        return

    full_df = pd.concat(uploaded_frames, ignore_index=True)

    # 2) Per-semester Faculty Presentables
    try:
        colmap = _map_columns(full_df)
        df_norm = _normalize_df(full_df, colmap)
        sem_col = colmap["Semester"]
        semesters_all = _semester_list(df_norm, sem_col)
        for sem in semesters_all:
            target_folder = BASE_REPORTS / f"{sanitize_sheet_name(sem)}_Dataset"
            build_faculty_presentables_for_semester(df_norm, sem, target_folder)
    except Exception as e:
        messagebox.showwarning("Faculty_Presentables", f"Skipped per-semester presentables:\n{e}")

    # 3) Faculty_Over_Time
    try:
        build_faculty_over_time_workbook(full_df, BASE_REPORTS)
    except Exception as e:
        messagebox.showwarning("Faculty_Over_Time", f"Skipped Faculty_Over_Time.xlsx:\n{e}")

    # 4) Semester_Over_Time
    try:
        build_semester_over_time_workbook(full_df, BASE_REPORTS, split_ws_ss=False)
    except Exception as e:
        messagebox.showwarning("Semester_Over_Time", f"Skipped Semester_Over_Time.xlsx:\n{e}")

    # 5) Semester_Faculty_Stats
    try:
        build_semester_faculty_stats_v2(full_df, BASE_REPORTS)
    except Exception as e:
        messagebox.showwarning("Semester_Faculty_Stats", f"Skipped Semester_Faculty_Stats.xlsx:\n{e}")

    # 6) Finish
    try:
        messagebox.showinfo(
            "Done",
            "Built reports in:\n"
            f"{BASE_REPORTS}\n\n"
            "Key files:\n"
            f"• {BASE_REPORTS / OUT_SEM_FAC_STATS}\n"
            f"• {BASE_REPORTS / OUT_FAC_OVER_TIME}\n"
            f"• {BASE_REPORTS / OUT_SEM_OVER_TIME}\n"
            "Plus one 'Faculty_Presentables.xlsx' per semester under <Semester>_Dataset/"
        )
    except Exception:
        print("Built reports in:", BASE_REPORTS)
        print(" -", OUT_SEM_FAC_STATS)
        print(" -", OUT_FAC_OVER_TIME)
        print(" -", OUT_SEM_OVER_TIME)
        print(" - And per-semester <Semester>_Dataset/Faculty_Presentables.xlsx folders")

if __name__ == "__main__":
    main()

# -------------------- Tiny headless test (optional) --------------------
def _tiny_headless_test():
    """
    Small synthetic dataset (3 faculties x 3 semesters)
    to exercise the Semester_Faculty_Stats builder + chart.
    NOT auto-run.
    """
    import numpy as np
    data = []
    faculties = [
        "01 Fakultät für Agrarwissenschaften",
        "Archiv07 Juristische Fakultät",
        "Lehreinheit Informatik (Fakultät für Mathematik und Informatik"
    ]
    semesters = ["SS 23", "WS 23/24", "SS 24"]
    rng = np.random.default_rng(42)
    for sem in semesters:
        for fac in faculties:
            n = rng.integers(30, 61)
            scores = rng.normal(loc=rng.uniform(65, 90), scale=8, size=n)
            for sc in scores:
                data.append(
                    {"Fakultät": fac, "Semester": sem, "Gesamtzufriedenheitsindex": float(np.clip(sc, 0, 100))}
                )
    test_df = pd.DataFrame(data)
    test_base = BASE_REPORTS / "TestRun"
    test_base.mkdir(parents=True, exist_ok=True)
    out_path = build_semester_faculty_stats_v2(test_df, test_base)
    print("[Test] Wrote:", out_path)

# To run the test instead of the GUI main, comment out the main() call above and call:
# _tiny_headless_test()
